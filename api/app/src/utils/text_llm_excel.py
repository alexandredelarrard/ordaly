"""
Build a styled multi-sheet XLSX from ``text_llm_by_schema`` (LLM extraction payload).

Sheets (see ``EXCEL_WORKBOOK_SHEETS``): ``summary``, ``offer_pictures``, ``rent_roll``,
``financial_statement``, ``building_report``, ``demographics_report``.

When ``source_pdf_path`` is set, each sheet embeds **key PDF pages** from
``page_extraction`` (see ``EXCEL_SHEET_PAGE_OF_INTEREST_FIELDS``) **below** the main
tables / content, at a **larger** resolution than the legacy offer-picture preset
(see ``_EMBED_MAX_WIDTH_PX`` / ``_EMBED_RENDER_ZOOM``).
"""

from __future__ import annotations

import io
import logging
import re
from datetime import date, datetime
from enum import Enum
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.utils.pdf_mupdf import open_pdf as _open_pdf_sanitized

from src.constants.variables import (
    EXCEL_COLUMN_ID_KEYS,
    EXCEL_FIELD_LABELS,
    EXCEL_SHEET_PAGE_OF_INTEREST_FIELDS,
    EXCEL_WORKBOOK_SHEETS,
    FINANCIAL_EXPENSE_ROWS,
    FINANCIAL_REVENUE_ROWS,
    FINANCIAL_ROW_ORDER,
    FINANCIAL_TOTAL_ROWS,
    RENT_ROLL_BLOCKS,
)

logger = logging.getLogger(__name__)

# Theme
_ACCENT = "0D3B66"
_WHITE = PatternFill("solid", fgColor="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor=_ACCENT)
_PROVENANCE_FILL = PatternFill("solid", fgColor="F7F9FC")
_SECTION_FILL = PatternFill("solid", fgColor="E8EEF4")
_REVENUE_FILL = PatternFill("solid", fgColor="E8F5E9")
_EXPENSE_FILL = PatternFill("solid", fgColor="FFF8E1")
_TOTAL_FILL = PatternFill("solid", fgColor="E3F2FD")
_DATA_ZEBRA = PatternFill("solid", fgColor="FAFBFC")

_TITLE_FONT = Font(name="Calibri", size=18, bold=True, color=_ACCENT)
_SUBTITLE_FONT = Font(name="Calibri", size=12, italic=True, color="5C6A7A")
_KPI_STRIP_FILL = PatternFill("solid", fgColor="F0F4F8")
_KPI_ACCENT_FILL = PatternFill("solid", fgColor="D4E4F4")

_VACANT_FONT = Font(name="Calibri", size=12, color="C00000")
_VACANT_FONT_BOLD = Font(name="Calibri", size=12, bold=True, color="C00000")

_HEADER_FONT = Font(name="Calibri", color="FFFFFF", bold=True, size=12)
_SECTION_FONT = Font(name="Calibri", bold=True, size=12, color=_ACCENT)
_BODY_FONT = Font(name="Calibri", size=12, color="1A1A1A")
_BODY_BOLD = Font(name="Calibri", size=12, bold=True, color="1A1A1A")
_TOTAL_FONT = Font(name="Calibri", size=12, bold=True, color=_ACCENT)

# LLM extraction confidence (0-10); green high, red low, black mid band.
_CONF_SCORE_GREEN = "008000"
_CONF_SCORE_RED = "C00000"
_CONF_SCORE_BLACK = "1A1A1A"
_THIN = Side(style="thin", color="D0D7DE")
_MEDIUM = Side(style="medium", color="9EB4C8")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_TOTAL_BORDER = Border(left=_THIN, right=_THIN, top=_MEDIUM, bottom=_MEDIUM)
_WRAP = Alignment(wrapText=True, vertical="top")
_LEFT = Alignment(vertical="top", horizontal="left")
_CENTER = Alignment(vertical="center", horizontal="center")
_RIGHT = Alignment(vertical="center", horizontal="right")

_FMT_CURRENCY = '"$"#,##0.00'
_FMT_CURRENCY_INT = '"$"#,##0'
_FMT_PERCENT = "0.00%"
_FMT_DATE = "mm/dd/yyyy"
_FMT_NUMBER = "#,##0.##"

_SHEET_WHITE_ROWS = 120
_SHEET_WHITE_COLS = 20
# Default cap for scaled PNG width (Excel display pixels). Used when no override.
_PICTURE_MAX_WIDTH_PX = 1080
# Sharper / wider raster for “key pages” blocks (all sheets + offer pictures).
_EMBED_MAX_WIDTH_PX = 1440
_EMBED_RENDER_ZOOM = 2.85
_PDF_RENDER_ZOOM = 2.35

_METAINFO_KEY = "page_number"
_YEAR_LABEL_KEYS = frozenset({"kpi_year", "revenue_year", "year", "year_label", "financial_year"})
_HOTEL_TOKENS = frozenset(
    ("hotel", "hospitality", "lodging", "resort", "motel", "inn", "hostel", "casino")
)

_PERCENT_RE = re.compile(
    r"(^|_)(cap_rate|occupancy|percentage|percent|_pct|growth_percentage|revpar_change)(_|$)",
    re.I,
)
_DATE_RE = re.compile(r"date", re.I)
# Avoid matching ``rent`` inside ``rentable``; avoid ``population`` inside ``total_population``.
_CURRENCY_RE = re.compile(
    r"(price|(?<![a-z])rent(?![a-z])|gross_potential|income|revenue|expense|fee|tax|noi|egi|gpr|adr|revpar|bid|reserve|"
    r"household_income|operating|profit|ebitda|miscellaneous|departmental|"
    r"undistributed|replacement|utilities|insurance|electric|gas|trash|water)",
    re.I,
)

# Explicit kinds (overrides substring heuristics on field names).
_FORCE_INTEGER_KEYS = frozenset(
    {
        "offer_rentable_square_feet",
        "unit_size_sf",
        "building_surface_sf",
        "amenity_size",
        "total_population",
        "total_households",
        "number_tenants_in_building",
        "buildings_number",
        "loading_dock_number",
        "floors_count",
        "parking_spaces",
        "building_height",
        "year_built",
        "year_renovated",
        "number_of_properties",
        "rooms_count",
        "beds_count",
        "lease_remaining_years",
    }
)
_FORCE_PERCENT_KEYS = frozenset(
    {
        "population_growth_projection",
        "occupancy_percentage",
    }
)
_DATE_PARSE_FORMATS = (
    "%Y-%m-%d",
    "%m/%d/%Y",
    "%m-%d-%Y",
    "%Y/%m/%d",
    "%d/%m/%Y",
    "%m/%d/%y",
    "%b %d, %Y",
    "%B %d, %Y",
)
# Control chars illegal in XLSX/XML (openpyxl); tab/newline/CR are allowed.
_ILLEGAL_XLSX_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _excel_value(value: Any) -> Any:
    """Strip characters that openpyxl rejects (common in PDF-extracted text)."""
    if value is None:
        return value
    if isinstance(value, str):
        return _ILLEGAL_XLSX_CHARS.sub("", value)
    if isinstance(value, (datetime, date, int, float, bool)):
        return value
    return _ILLEGAL_XLSX_CHARS.sub("", str(value))


class CellKind(str, Enum):
    TEXT = "text"
    CURRENCY = "currency"
    PERCENT = "percent"
    DATE = "date"
    INTEGER = "integer"
    NUMBER = "number"


# ---------------------------------------------------------------------------
# Value helpers
# ---------------------------------------------------------------------------


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return not value.strip()
    if isinstance(value, (list, dict)):
        return len(value) == 0
    return False


def _has_meaningful_value(value: Any) -> bool:
    if isinstance(value, dict):
        return any(_has_meaningful_value(v) for k, v in value.items() if k != _METAINFO_KEY)
    if isinstance(value, list):
        return any(_has_meaningful_value(v) for v in value)
    return not _is_blank(value)


def _dict_all_values_blank(d: dict[str, Any]) -> bool:
    return all(_is_blank(v) for v in d.values())


def _flatten_item(item: dict[str, Any], *, depth: int = 0) -> dict[str, Any]:
    flat: dict[str, Any] = {}
    for key, value in item.items():
        if key == _METAINFO_KEY:
            continue
        if isinstance(value, dict):
            if "in_place_t12" in value or "pro_forma_year1" in value:
                for period in ("in_place_t12", "pro_forma_year1"):
                    if value.get(period) is not None:
                        flat[f"{key}.{period}"] = value[period]
            elif depth < 2:
                for sub_k, sub_v in value.items():
                    flat[f"{key}.{sub_k}"] = sub_v
            else:
                flat[key] = value
        else:
            flat[key] = value
    return flat


def _filter_list_columns(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for raw in items:
        if not isinstance(raw, dict):
            continue
        flat = _flatten_item(raw)
        if flat and not _dict_all_values_blank(flat):
            out.append(flat)
    return out


def _infer_cell_kind(field_key: str, value: Any) -> CellKind:
    if field_key in _YEAR_LABEL_KEYS:
        return CellKind.TEXT
    if field_key in _FORCE_PERCENT_KEYS:
        return CellKind.PERCENT
    if field_key in _FORCE_INTEGER_KEYS:
        if isinstance(value, bool):
            return CellKind.TEXT
        if isinstance(value, int):
            return CellKind.INTEGER
        if isinstance(value, float):
            return CellKind.INTEGER if value == int(value) else CellKind.NUMBER
        if isinstance(value, str):
            n = _coerce_numeric(value)
            if n is None:
                return CellKind.TEXT
            if isinstance(n, float) and n != int(n):
                return CellKind.NUMBER
            return CellKind.INTEGER
        return CellKind.TEXT
    if _DATE_RE.search(field_key):
        return CellKind.DATE
    if _PERCENT_RE.search(field_key):
        return CellKind.PERCENT
    if isinstance(value, bool):
        return CellKind.TEXT
    if isinstance(value, int) and not _CURRENCY_RE.search(field_key):
        if "year" in field_key.lower() and "founded" not in field_key.lower():
            return CellKind.INTEGER
        if _CURRENCY_RE.search(field_key):
            return CellKind.CURRENCY
        return CellKind.INTEGER
    if isinstance(value, float):
        if _PERCENT_RE.search(field_key):
            return CellKind.PERCENT
        if _CURRENCY_RE.search(field_key):
            return CellKind.CURRENCY
        return CellKind.NUMBER
    if isinstance(value, str) and _parse_date(value) is not None:
        return CellKind.DATE
    return CellKind.TEXT


def _parse_date(value: Any) -> date | datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return value
    if not isinstance(value, str):
        return None
    text = _excel_value(value).strip()
    if not text:
        return None
    for fmt in _DATE_PARSE_FORMATS:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _coerce_numeric(value: Any) -> float | int | None:
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        cleaned = _excel_value(value).strip().replace(",", "").replace("$", "").replace("%", "")
        if not cleaned:
            return None
        try:
            num = float(cleaned)
            return int(num) if num == int(num) else num
        except ValueError:
            return None
    return None


# Negative Total / Other OpEx are omitted in the financial matrix (see ``_write_list_matrix``).
_FINANCIAL_SUPPRESS_NEGATIVE_OPEX_KEYS = frozenset(
    {"total_operating_expenses", "other_operating_expenses"}
)


def _all_numeric_values_negative_for_key(columns: list[dict[str, Any]], key: str) -> bool:
    """True when at least one column has a number and every such number is < 0."""
    found = False
    for col in columns:
        n = _coerce_numeric(col.get(key))
        if n is None:
            continue
        found = True
        if n >= 0:
            return False
    return found


def _prepare_cell(field_key: str, value: Any) -> tuple[Any, CellKind]:
    if _is_blank(value):
        return "", CellKind.TEXT

    kind = _infer_cell_kind(field_key, value)

    if kind == CellKind.DATE:
        parsed = _parse_date(value)
        return (parsed if parsed is not None else value), CellKind.DATE

    if kind == CellKind.PERCENT:
        num = _coerce_numeric(value)
        if num is None:
            return value, CellKind.TEXT
        # Percent points (e.g. 6.5, 75) vs fraction (0.065); growth often 1.2–5.
        if abs(num) > 1.5 or (field_key in _FORCE_PERCENT_KEYS and 1 < abs(num) <= 100):
            num = num / 100.0
        return num, CellKind.PERCENT

    if kind in (CellKind.CURRENCY, CellKind.INTEGER, CellKind.NUMBER):
        num = _coerce_numeric(value)
        if num is None:
            return value, CellKind.TEXT
        return num, kind

    if isinstance(value, bool):
        return ("Yes" if value else "No"), CellKind.TEXT

    if isinstance(value, float) and value == int(value):
        return int(value), CellKind.INTEGER

    if isinstance(value, str):
        return _excel_value(value), CellKind.TEXT

    return value, CellKind.TEXT


def _unit_is_vacant(unit: dict[str, Any]) -> bool:
    status = str(unit.get("unit_rent_status") or "").strip().lower()
    if "vacant" in status:
        return True
    tenant = str(unit.get("tenant_name") or "").strip().lower()
    return tenant in ("vacant", "empty", "available", "-")


def _apply_cell_style(
    cell: Any,
    *,
    kind: CellKind,
    row_fill: PatternFill | None = None,
    bold_label: bool = False,
    bold_value: bool = False,
    is_total_row: bool = False,
    vacant: bool = False,
) -> None:
    cell.border = _TOTAL_BORDER if is_total_row else _BORDER
    cell.alignment = _RIGHT if kind in (
        CellKind.CURRENCY,
        CellKind.PERCENT,
        CellKind.INTEGER,
        CellKind.NUMBER,
    ) else _WRAP
    if kind == CellKind.CURRENCY:
        v = cell.value
        if isinstance(v, (int, float)) and v == int(v):
            cell.number_format = _FMT_CURRENCY_INT
        else:
            cell.number_format = _FMT_CURRENCY
    elif kind == CellKind.PERCENT:
        cell.number_format = _FMT_PERCENT
    elif kind == CellKind.DATE:
        cell.number_format = _FMT_DATE
    elif kind == CellKind.INTEGER:
        cell.number_format = "#,##0"
    elif kind == CellKind.NUMBER:
        cell.number_format = _FMT_NUMBER

    if vacant and not is_total_row:
        cell.font = _VACANT_FONT_BOLD if (bold_value or bold_label) else _VACANT_FONT
        cell.fill = row_fill or _WHITE
    elif is_total_row:
        cell.font = _TOTAL_FONT
        cell.fill = _TOTAL_FILL
    elif bold_value:
        cell.font = _BODY_BOLD
        cell.fill = row_fill or _WHITE
    else:
        cell.font = _BODY_FONT
        cell.fill = row_fill or _WHITE

    if bold_label and cell.column == 1:
        if vacant and not is_total_row:
            cell.font = _VACANT_FONT_BOLD
        else:
            cell.font = _TOTAL_FONT if is_total_row else _BODY_BOLD
        cell.alignment = _LEFT


def _ordered_row_keys(
    columns: list[dict[str, Any]],
    label_section: str,
    *,
    custom_order: tuple[str, ...] | None = None,
) -> list[str]:
    skip = EXCEL_COLUMN_ID_KEYS.get(label_section, frozenset())
    seen: set[str] = set()
    keys: list[str] = []
    order = custom_order or tuple(EXCEL_FIELD_LABELS.get(label_section, {}).keys())
    for k in order:
        if k in skip or k in seen:
            continue
        if any(k in col for col in columns):
            keys.append(k)
            seen.add(k)
    for col in columns:
        for k in col:
            if k in skip or k in seen:
                continue
            keys.append(k)
            seen.add(k)
    return keys


def _row_label(label_section: str, field_key: str) -> str:
    return EXCEL_FIELD_LABELS.get(label_section, {}).get(
        field_key,
        EXCEL_FIELD_LABELS.get("_common", {}).get(
            field_key, field_key.replace("_", " ").replace(".", " — ").title()
        ),
    )


def _collect_page_numbers(*sections: Any) -> str:
    parts: list[str] = []
    seen: set[str] = set()

    def _add(raw: Any) -> None:
        if raw is None:
            return
        text = str(raw).strip()
        if not text or text in seen:
            return
        seen.add(text)
        parts.append(text)

    for section in sections:
        if isinstance(section, dict):
            _add(section.get(_METAINFO_KEY))
        elif isinstance(section, list):
            for item in section:
                if isinstance(item, dict):
                    _add(item.get(_METAINFO_KEY))
    return ", ".join(parts)


def _page_extraction_dict(data: dict[str, Any]) -> dict[str, Any]:
    raw = data.get("page_extraction")
    return raw if isinstance(raw, dict) else {}


def _normalize_page_numbers(raw: Any) -> list[int]:
    """1-based PDF page indices from ``PageOfInterest`` list fields."""
    if raw is None:
        return []
    items = raw if isinstance(raw, list) else [raw]
    out: list[int] = []
    seen: set[int] = set()
    for item in items:
        n = _coerce_numeric(item)
        if n is None:
            continue
        page = int(n)
        if page < 1 or page in seen:
            continue
        seen.add(page)
        out.append(page)
    return sorted(out)


def _format_page_list(pages: list[int]) -> str:
    if not pages:
        return ""
    return ", ".join(str(p) for p in pages)


def _page_list_for_poi_fields(data: dict[str, Any], fields: str | tuple[str, ...]) -> list[int]:
    poi = _page_extraction_dict(data)
    names = (fields,) if isinstance(fields, str) else fields
    merged: list[int] = []
    seen: set[int] = set()
    for name in names:
        for p in _normalize_page_numbers(poi.get(name)):
            if p not in seen:
                seen.add(p)
                merged.append(p)
    return sorted(merged)


def _pages_for_sheet(data: dict[str, Any], sheet_name: str, *legacy_sections: Any) -> str:
    """Pages (OM) from ``page_extraction``, merged with any legacy ``page_number`` on sections."""
    parts: list[str] = []
    seen: set[str] = set()

    def _add_text(text: str) -> None:
        t = text.strip()
        if not t or t in seen:
            return
        seen.add(t)
        parts.append(t)

    poi_fields = EXCEL_SHEET_PAGE_OF_INTEREST_FIELDS.get(sheet_name)
    if poi_fields:
        _add_text(_format_page_list(_page_list_for_poi_fields(data, poi_fields)))

    legacy = _collect_page_numbers(*legacy_sections)
    if legacy:
        for token in legacy.split(","):
            _add_text(token.strip())

    return ", ".join(parts)


def _render_pdf_page_png(
    pdf_path: Path,
    page_num_1based: int,
    *,
    zoom: float | None = None,
) -> bytes | None:
    """Rasterize one PDF page to PNG bytes (``page_num_1based`` is 1-based)."""
    try:
        import pymupdf
    except ImportError:
        logger.warning("pymupdf not installed — cannot render offer pictures.")
        return None

    z = float(zoom) if zoom is not None else float(_PDF_RENDER_ZOOM)

    try:
        doc = _open_pdf_sanitized(pdf_path)
        try:
            idx = page_num_1based - 1
            if idx < 0 or idx >= doc.page_count:
                return None
            page = doc[idx]
            mat = pymupdf.Matrix(z, z)
            pix = page.get_pixmap(matrix=mat, alpha=False)
            return pix.tobytes("png")
        finally:
            doc.close()
    except Exception as exc:
        logger.warning("Failed to render PDF page %s from %s: %s", page_num_1based, pdf_path, exc)
        return None


def _scaled_xl_image(
    png_bytes: bytes,
    *,
    max_width_px: int | None = None,
) -> XLImage | None:
    try:
        from PIL import Image as PILImage
    except ImportError:
        logger.warning("Pillow not installed — cannot embed offer pictures.")
        return None

    cap = int(max_width_px) if max_width_px is not None else int(_PICTURE_MAX_WIDTH_PX)
    if cap < 1:
        cap = int(_PICTURE_MAX_WIDTH_PX)

    pil = PILImage.open(io.BytesIO(png_bytes))
    w, h = pil.size
    if w <= 0 or h <= 0:
        return None
    scale = min(1.0, cap / w)
    disp_w = max(int(w * scale), 1)
    disp_h = max(int(h * scale), 1)
    buf = io.BytesIO()
    resample = getattr(getattr(PILImage, "Resampling", PILImage), "LANCZOS", PILImage.LANCZOS)
    pil.resize((disp_w, disp_h), resample).save(buf, format="PNG")
    buf.seek(0)
    img = XLImage(buf)
    img.width = disp_w
    img.height = disp_h
    return img


def _anchor_image_rows(image_height_px: int) -> int:
    """Approximate Excel row span for a floating image."""
    return max(int(image_height_px / 15) + 3, 6)


def _embed_pdf_screenshots_block(
    ws: Any,
    row: int,
    data: dict[str, Any],
    *,
    poi_fields: str | tuple[str, ...],
    source_pdf_path: Path | None,
    banner_title: str,
    render_zoom: float | None = None,
    max_width_px: int | None = None,
) -> int:
    """
    Append rasterized PDF pages (from ``page_extraction`` POI fields) under ``row``.

    Uses ``render_zoom`` / ``max_width_px`` when set; otherwise ``_EMBED_RENDER_ZOOM`` /
    ``_EMBED_MAX_WIDTH_PX`` for larger, sharper sheet screenshots than the legacy cap.
    """
    pages = _page_list_for_poi_fields(data, poi_fields)
    pdf_ok = source_pdf_path is not None and Path(source_pdf_path).is_file()
    if not pages:
        return row
    if not pdf_ok:
        logger.warning("%s: skipping PDF screenshots (path missing or not found).", banner_title)
        return row

    z = float(render_zoom) if render_zoom is not None else float(_EMBED_RENDER_ZOOM)
    cap = int(max_width_px) if max_width_px is not None else int(_EMBED_MAX_WIDTH_PX)
    if cap < 1:
        cap = int(_EMBED_MAX_WIDTH_PX)

    pdf_path = Path(source_pdf_path).resolve()
    cur_w = ws.column_dimensions["A"].width
    try:
        base_w = float(cur_w) if cur_w is not None else 20.0
    except (TypeError, ValueError):
        base_w = 20.0
    # Wider column so large images are usable in Excel (width is in “character units”).
    ws.column_dimensions["A"].width = max(base_w, min(120.0, 18.0 + cap / 90.0))

    row += 1
    last_col = max(int(ws.max_column or 1), 6)
    row = _write_section_banner(ws, row, banner_title, last_col)

    embedded = 0
    for page_num in pages:
        png = _render_pdf_page_png(pdf_path, page_num, zoom=z)
        if not png:
            continue
        xl_img = _scaled_xl_image(png, max_width_px=cap)
        if xl_img is None:
            continue

        ws.cell(row=row, column=1, value=f"Page {page_num}").font = _SECTION_FONT
        ws.cell(row=row, column=1).fill = _SECTION_FILL
        ws.cell(row=row, column=1).border = _BORDER
        row += 1

        ws.add_image(xl_img, f"A{row}")
        row += _anchor_image_rows(int(xl_img.height))
        embedded += 1
        row += 1

    if embedded == 0:
        miss = ws.cell(
            row=row,
            column=1,
            value="Could not render PDF pages for this section (check PyMuPDF / Pillow).",
        )
        miss.font = _BODY_FONT
        row += 1

    return row


# Workbook tab name (``EXCEL_WORKBOOK_SHEETS``) → section banner for PDF screenshots.
_EXCEL_KEY_PAGES_BANNER: dict[str, str] = {
    "summary": "Summary — key PDF pages (screenshots)",
    "offer_pictures": "Property pictures — key PDF pages (screenshots)",
    "rent_roll": "Rent roll — key PDF pages (screenshots)",
    "financial_statement": "Financial statement — key PDF pages (screenshots)",
    "building_report": "Building report — key PDF pages (screenshots)",
    "demographics_report": "Demographics & attractiveness — key PDF pages (screenshots)",
}


def _embed_excel_sheet_key_pages_below(
    ws: Any,
    row: int,
    data: dict[str, Any],
    sheet_key: str,
    source_pdf_path: Path | None,
) -> int:
    """
    Append screenshots for the ``PageOfInterest`` field(s) mapped to this Excel tab
    (see ``EXCEL_SHEET_PAGE_OF_INTEREST_FIELDS``), using embed zoom / max width.
    """
    poi = EXCEL_SHEET_PAGE_OF_INTEREST_FIELDS.get(sheet_key)
    if not poi:
        return row
    banner = _EXCEL_KEY_PAGES_BANNER.get(
        sheet_key,
        f"{sheet_key.replace('_', ' ').title()} — key PDF pages (screenshots)",
    )
    return _embed_pdf_screenshots_block(
        ws,
        row,
        data,
        poi_fields=poi,
        source_pdf_path=source_pdf_path,
        banner_title=banner,
    )


# Keys stored on each ``Leases`` item (schema); repeated unit/rent fields come from ``RentRollRow``.
_RENT_ROLL_LEASE_ITEM_KEYS: frozenset[str] = frozenset(
    {
        "lease_start_date",
        "lease_end_date",
        "lease_remaining_years",
        "lease_structure_type",
    }
)


def _iter_rent_roll_excel_rows(unit: dict[str, Any]) -> list[dict[str, Any]]:
    """
    Build one flat dict per **Excel row**: unit/rent/escalation fields repeat; each
    ``Leases`` entry becomes its own row. Legacy single ``unit_leases`` dict is one row.

    ``unit_leases`` absent or ``[]`` yields a single row (lease columns may be empty).
    """
    skip = frozenset({"unit_leases", _METAINFO_KEY})
    base: dict[str, Any] = {k: v for k, v in unit.items() if k not in skip}
    ul = unit.get("unit_leases")

    def _row_from_lease(lease: dict[str, Any]) -> dict[str, Any]:
        row = dict(base)
        for lk in _RENT_ROLL_LEASE_ITEM_KEYS:
            row.pop(lk, None)
        for lk in _RENT_ROLL_LEASE_ITEM_KEYS:
            v = lease.get(lk)
            if not _is_blank(v):
                row[lk] = v
        return row

    if isinstance(ul, list):
        lease_rows = [_row_from_lease(le) for le in ul if isinstance(le, dict)]
        return lease_rows if lease_rows else [dict(base)]

    if isinstance(ul, dict) and ul:
        return [_row_from_lease(ul)]

    return [dict(base)]


def _rent_roll_building_sections(rr: dict[str, Any]) -> list[tuple[str | None, list[dict[str, Any]]]]:
    """
    ``RentRollReport.rows`` is either a list of ``RentRollReportPerBuilding`` dicts
    (each with ``building_name`` + nested ``rows``), or a legacy flat list of units.
    """
    outer = rr.get("rows")
    if not isinstance(outer, list) or not outer:
        return []
    first = outer[0]
    if isinstance(first, dict) and isinstance(first.get("rows"), list):
        sections: list[tuple[str | None, list[dict[str, Any]]]] = []
        for b in outer:
            if not isinstance(b, dict):
                continue
            name = b.get("building_name")
            inner = b.get("rows") or []
            units = [x for x in inner if isinstance(x, dict)]
            label = str(name).strip() if not _is_blank(name) else None
            sections.append((label, units))
        return sections
    label = str(rr.get("building_name") or "").strip() or None
    return [
        (
            label,
            [x for x in outer if isinstance(x, dict)],
        )
    ]


def _first_financial_year_kpis(
    data: dict[str, Any],
) -> tuple[Any, CellKind, Any, CellKind, str | None]:
    """Cap rate, NOI, and period label from the first row of ``financial_cycles``."""
    fin_key = _pick_financial_key(data)
    if not fin_key:
        return "", CellKind.TEXT, "", CellKind.TEXT, None
    section = data.get(fin_key)
    if not isinstance(section, dict):
        return "", CellKind.TEXT, "", CellKind.TEXT, None
    cycles = section.get("financial_cycles")
    if not isinstance(cycles, list) or not cycles:
        return "", CellKind.TEXT, "", CellKind.TEXT, None
    first = cycles[0]
    if not isinstance(first, dict):
        return "", CellKind.TEXT, "", CellKind.TEXT, None
    year = first.get("financial_year") or first.get("year_label") or first.get("year")
    year_s = str(year).strip() if not _is_blank(year) else None
    cap_raw = first.get("cap_rate")
    noi_raw = first.get("net_operating_income")
    cap_disp, cap_kind = _prepare_cell("cap_rate", cap_raw)
    noi_disp, noi_kind = _prepare_cell("net_operating_income", noi_raw)
    return cap_disp, cap_kind, noi_disp, noi_kind, year_s


def _metadata_scalar_block(meta: dict[str, Any]) -> dict[str, Any]:
    """Top-level metadata fields excluding nested ``asset`` list."""
    skip = frozenset({"asset", _METAINFO_KEY})
    return {k: v for k, v in meta.items() if k not in skip and not _is_blank(v)}


def _metadata_offer_rows(meta: dict[str, Any]) -> list[dict[str, Any]]:
    raw = meta.get("asset")
    if not isinstance(raw, list):
        return []
    return [x for x in raw if isinstance(x, dict)]


def _write_hero_header(ws: Any, row: int, title: str, subtitle: str | None, last_col: int) -> int:
    lc = max(last_col, 4)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=lc)
    cell = ws.cell(row=row, column=1, value=_excel_value(title))
    cell.font = _TITLE_FONT
    cell.fill = _WHITE
    cell.alignment = Alignment(wrapText=True, vertical="center", horizontal="left")
    cell.border = Border(bottom=_MEDIUM)
    r = row + 1
    if subtitle and subtitle.strip():
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=lc)
        sub = ws.cell(row=r, column=1, value=_excel_value(subtitle))
        sub.font = _SUBTITLE_FONT
        sub.fill = _WHITE
        sub.alignment = _WRAP
        sub.border = _BORDER
        r += 1
    return r + 1


def _write_financial_kpi_strip(
    ws: Any,
    start_row: int,
    *,
    year_label: str | None,
    cap_val: Any,
    cap_kind: CellKind,
    noi_val: Any,
    noi_kind: CellKind,
) -> int:
    """Two-column KPI layout: cap rate and NOI from the first financial year."""
    r = start_row
    suffix = f" — {year_label}" if year_label else ""
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    banner = ws.cell(row=r, column=1, value=_excel_value(f"Key financial indicators{suffix}"))
    banner.font = _SECTION_FONT
    banner.fill = _KPI_STRIP_FILL
    banner.alignment = _LEFT
    banner.border = _BORDER
    r += 1

    has_cap = not _is_blank(cap_val) and cap_val != ""
    has_noi = not _is_blank(noi_val) and noi_val != ""

    if not has_cap and not has_noi:
        ws.cell(row=r, column=1, value="No cap rate or NOI on the first financial period.").font = _BODY_FONT
        ws.cell(row=r, column=1).fill = _WHITE
        return r + 2

    # Row: [Cap label][Cap value][spacer][NOI label][NOI value]
    label_cap = ws.cell(row=r, column=1, value="Cap rate")
    label_cap.font = _BODY_BOLD
    label_cap.fill = _KPI_ACCENT_FILL
    label_cap.border = _BORDER
    label_cap.alignment = _LEFT
    v_cap = ws.cell(row=r, column=2, value=_excel_value(cap_val) if has_cap else "—")
    _apply_cell_style(v_cap, kind=cap_kind if has_cap else CellKind.TEXT, row_fill=_WHITE, bold_value=True)

    label_noi = ws.cell(row=r, column=4, value="NOI")
    label_noi.font = _BODY_BOLD
    label_noi.fill = _KPI_ACCENT_FILL
    label_noi.border = _BORDER
    label_noi.alignment = _LEFT
    v_noi = ws.cell(row=r, column=5, value=_excel_value(noi_val) if has_noi else "—")
    _apply_cell_style(v_noi, kind=noi_kind if has_noi else CellKind.TEXT, row_fill=_WHITE, bold_value=True)
    ws.cell(row=r, column=3, value="").fill = _WHITE
    ws.cell(row=r, column=6, value="").fill = _WHITE
    return r + 2


def _offer_line_column_titles(columns: list[dict[str, Any]], data: dict[str, Any]) -> list[str]:
    """Column headers for the summary offer matrix: prefer ``building_name`` from OM data."""
    n = len(columns)
    titles: list[str | None] = [None] * n

    for i in range(n):
        col = columns[i]
        if isinstance(col, dict):
            bn = col.get("building_name")
            if not _is_blank(bn):
                titles[i] = str(bn).strip()

    br = data.get("building_report")
    assets = br.get("assets") if isinstance(br, dict) else None
    if isinstance(assets, list):
        for i in range(min(n, len(assets))):
            if isinstance(assets[i], dict):
                bn = assets[i].get("building_name")
                if not _is_blank(bn):
                    titles[i] = str(bn).strip()

    rr = data.get("rent_roll_report")
    rr_rows = rr.get("rows") if isinstance(rr, dict) and isinstance(rr.get("rows"), list) else []
    rr_names: list[str] = []
    for sec in rr_rows:
        if isinstance(sec, dict) and isinstance(sec.get("rows"), list):
            bn = sec.get("building_name")
            if not _is_blank(bn):
                rr_names.append(str(bn).strip())
    for i in range(n):
        if titles[i] is None and i < len(rr_names):
            titles[i] = rr_names[i]

    out: list[str] = []
    for i in range(n):
        label = titles[i] if titles[i] else f"Offer {i + 1}"
        out.append(_excel_value(str(label))[:40])
    return out


def _write_offer_lines_matrix(
    ws: Any,
    start_row: int,
    offers: list[dict[str, Any]],
    *,
    data: dict[str, Any] | None = None,
) -> int:
    """One column per priced offer / building bucket from metadata."""
    if not offers:
        return start_row
    columns = _filter_list_columns(offers)
    if not columns:
        return start_row
    order = tuple(EXCEL_FIELD_LABELS.get("offer_line", {}).keys())
    row_keys = _row_keys_for_allowed(columns, order)
    if not row_keys:
        row_keys = _ordered_row_keys(columns, "offer_line", custom_order=order or None)
    if not row_keys:
        return start_row
    last_col = 1 + len(columns)
    hdr_row = _write_section_banner(ws, start_row, "Offer line items", max(last_col, 2))
    _style_table_header(ws, hdr_row, 1, last_col)
    ws.cell(row=hdr_row, column=1, value="Metric")
    col_titles = _offer_line_column_titles(columns, data or {})
    if len(col_titles) != len(columns):
        col_titles = [f"Offer {j + 1}" for j in range(len(columns))]
    for j, title in enumerate(col_titles):
        hdr = ws.cell(row=hdr_row, column=j + 2, value=title)
        hdr.font = _HEADER_FONT
        hdr.fill = _HEADER_FILL
        hdr.border = _BORDER
        hdr.alignment = _CENTER
    r = hdr_row + 1
    for idx, key in enumerate(row_keys):
        row_fill = _DATA_ZEBRA if idx % 2 else _WHITE
        lbl = ws.cell(row=r, column=1, value=_excel_value(_row_label("offer_line", key)))
        _apply_cell_style(lbl, kind=CellKind.TEXT, row_fill=row_fill, bold_label=True)
        for j, col_data in enumerate(columns):
            raw = col_data.get(key)
            display, kind = _prepare_cell(key, raw)
            cell = ws.cell(row=r, column=j + 2, value=_excel_value(display))
            _apply_cell_style(cell, kind=kind, row_fill=row_fill)
        r += 1
    return r + 1


def _is_hotel_asset(data: dict[str, Any]) -> bool:
    meta = data.get("metadata_from_text")
    if not isinstance(meta, dict):
        return False
    asset = (meta.get("asset_type") or "").strip().lower()
    return any(tok in asset for tok in _HOTEL_TOKENS)


def _pick_financial_key(data: dict[str, Any]) -> str | None:
    hotel = data.get("financial_statement_hotel")
    standard = data.get("financial_statement")
    hotel_ok = _has_meaningful_value(hotel)
    standard_ok = _has_meaningful_value(standard)
    if hotel_ok and (not standard_ok or _is_hotel_asset(data)):
        return "financial_statement_hotel"
    if standard_ok:
        return "financial_statement"
    if hotel_ok:
        return "financial_statement_hotel"
    return None


def _financial_row_fill(fin_key: str, row_key: str) -> PatternFill | None:
    if row_key in FINANCIAL_TOTAL_ROWS.get(fin_key, frozenset()):
        return _TOTAL_FILL
    if row_key in FINANCIAL_REVENUE_ROWS.get(fin_key, frozenset()):
        return _REVENUE_FILL
    if row_key in FINANCIAL_EXPENSE_ROWS.get(fin_key, frozenset()):
        return _EXPENSE_FILL
    return None


def _financial_is_total(fin_key: str, row_key: str) -> bool:
    return row_key in FINANCIAL_TOTAL_ROWS.get(fin_key, frozenset())


# ---------------------------------------------------------------------------
# Sheet canvas
# ---------------------------------------------------------------------------


def _init_sheet_canvas(ws: Any) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = None
    for r in range(1, _SHEET_WHITE_ROWS + 1):
        for c in range(1, _SHEET_WHITE_COLS + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = _WHITE
            cell.font = _BODY_FONT


def _autosize_columns(ws: Any, max_col: int, min_width: int = 12, max_width: int = 44) -> None:
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        best = min_width
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value is not None and str(cell.value):
                    best = max(best, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[letter].width = best


def _style_table_header(ws: Any, row: int, col_start: int, col_end: int) -> None:
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _BORDER
        cell.alignment = _CENTER


def _parse_confidence_answer(section: dict[str, Any] | None) -> int | None:
    """Return clamped 0-10 score from ``confidence_answer`` on a schema dict, or None."""
    if not isinstance(section, dict):
        return None
    raw = section.get("confidence_answer")
    if raw is None or raw == "":
        return None
    try:
        n = int(float(raw))
    except (TypeError, ValueError):
        return None
    return max(0, min(10, n))


def _confidence_score_value_font(score: int | None) -> Font:
    if score is None:
        return _BODY_FONT
    if score >= 7:
        return Font(name="Calibri", size=12, bold=True, color=_CONF_SCORE_GREEN)
    if score <= 4:
        return Font(name="Calibri", size=12, bold=True, color=_CONF_SCORE_RED)
    return Font(name="Calibri", size=12, bold=True, color=_CONF_SCORE_BLACK)


def _write_page_block(
    ws: Any,
    row: int,
    pages: str,
    *,
    confidence_section: dict[str, Any] | None = None,
) -> int:
    ws.cell(row=row, column=1, value="Pages (OM)").font = _BODY_BOLD
    ws.cell(row=row, column=1).fill = _PROVENANCE_FILL
    ws.cell(row=row, column=1).border = _BORDER
    val_cell = ws.cell(row=row, column=2, value=_excel_value(pages or "—"))
    val_cell.font = _BODY_FONT
    val_cell.fill = _PROVENANCE_FILL
    val_cell.border = _BORDER
    val_cell.alignment = _WRAP

    score = _parse_confidence_answer(confidence_section)
    r2 = row + 1
    ws.cell(row=r2, column=1, value="Confidence (0-10)").font = _BODY_BOLD
    ws.cell(row=r2, column=1).fill = _PROVENANCE_FILL
    ws.cell(row=r2, column=1).border = _BORDER
    conf_val = ws.cell(row=r2, column=2, value="—" if score is None else score)
    conf_val.font = _confidence_score_value_font(score)
    conf_val.fill = _PROVENANCE_FILL
    conf_val.border = _BORDER
    conf_val.alignment = _WRAP
    return row + 3


def _write_section_banner(ws: Any, row: int, title: str, last_col: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(last_col, 2))
    cell = ws.cell(row=row, column=1, value=_excel_value(title))
    cell.font = _SECTION_FONT
    cell.fill = _SECTION_FILL
    cell.alignment = _LEFT
    cell.border = _BORDER
    return row + 1


def _write_single_column_table(
    ws: Any,
    start_row: int,
    data: dict[str, Any],
    *,
    label_section: str,
) -> int:
    rows = [(k, v) for k, v in data.items() if k != _METAINFO_KEY and not _is_blank(v)]
    if not rows:
        return start_row

    _style_table_header(ws, start_row, 1, 2)
    ws.cell(row=start_row, column=1, value="Field")
    ws.cell(row=start_row, column=2, value="Value")
    r = start_row + 1
    for idx, (key, raw_val) in enumerate(rows):
        display, kind = _prepare_cell(key, raw_val)
        label_cell = ws.cell(row=r, column=1, value=_excel_value(_row_label(label_section, key)))
        _apply_cell_style(label_cell, kind=CellKind.TEXT, row_fill=_DATA_ZEBRA if idx % 2 else _WHITE, bold_label=True)
        val_cell = ws.cell(row=r, column=2, value=_excel_value(display))
        _apply_cell_style(val_cell, kind=kind, row_fill=_DATA_ZEBRA if idx % 2 else _WHITE)
        r += 1
    return r + 1


def _column_header(item: dict[str, Any], index: int, label_section: str) -> str:
    """Build column title from schema id fields (year, unit name, area, etc.)."""
    if label_section == "rent_roll_report":
        parts: list[str] = []
        unit = item.get("unit_id") or item.get("unit_name")
        tenant = item.get("tenant_name")
        if not _is_blank(unit):
            parts.append(_excel_value(str(unit)).strip())
        if not _is_blank(tenant):
            parts.append(_excel_value(str(tenant)).strip())
        if parts:
            return _excel_value(" — ".join(parts)[:40])

    for key in EXCEL_COLUMN_ID_KEYS.get(label_section, frozenset()):
        val = item.get(key)
        if not _is_blank(val):
            return _excel_value(str(val).strip()[:40])

    for key in (
        "financial_year",
        "kpi_year",
        "revenue_year",
        "year",
        "year_label",
        "radius_scope",
        "radius_label",
        "unit_id",
        "unit_name",
        "tenant_name",
        "name",
    ):
        val = item.get(key)
        if not _is_blank(val):
            return _excel_value(str(val).strip()[:40])
    return f"Column {index + 1}"


def _row_keys_for_allowed(
    columns: list[dict[str, Any]],
    allowed_keys: tuple[str, ...],
) -> list[str]:
    keys: list[str] = []
    for k in allowed_keys:
        if any(not _is_blank(col.get(k)) for col in columns):
            keys.append(k)
    return keys


def _write_list_matrix(
    ws: Any,
    start_row: int,
    items: list[Any],
    *,
    label_section: str,
    custom_order: tuple[str, ...] | None = None,
    allowed_keys: tuple[str, ...] | None = None,
    default_row_fill: PatternFill | None = None,
    columns: list[dict[str, Any]] | None = None,
    row_fill_fn: Any = None,
    row_is_total_fn: Any = None,
    repeat_column_headers: bool = True,
    vacant_columns: list[bool] | None = None,
) -> int:
    if columns is None:
        columns = _filter_list_columns([x for x in items if isinstance(x, dict)])
    if not columns:
        return start_row

    if allowed_keys is not None:
        row_keys = _row_keys_for_allowed(columns, allowed_keys)
    else:
        row_keys = _ordered_row_keys(columns, label_section, custom_order=custom_order)
    if not row_keys:
        return start_row

    if label_section in ("financial_statement", "financial_statement_hotel"):
        row_keys = [
            k
            for k in row_keys
            if not (
                k in _FINANCIAL_SUPPRESS_NEGATIVE_OPEX_KEYS
                and _all_numeric_values_negative_for_key(columns, k)
            )
        ]
        if not row_keys:
            return start_row

    last_col = 1 + len(columns)
    header_row = start_row
    if repeat_column_headers:
        _style_table_header(ws, header_row, 1, last_col)
        ws.cell(row=header_row, column=1, value="Field")
        for j, col_data in enumerate(columns):
            is_vacant = bool(vacant_columns[j]) if vacant_columns else False
            hdr = ws.cell(
                row=header_row,
                column=j + 2,
                value=_excel_value(_column_header(col_data, j, label_section)),
            )
            if is_vacant:
                hdr.font = _VACANT_FONT_BOLD
                hdr.fill = PatternFill("solid", fgColor="FDECEA")
            else:
                hdr.font = _HEADER_FONT
                hdr.fill = _HEADER_FILL
            hdr.border = _BORDER
            hdr.alignment = _CENTER
        data_start = header_row + 1
    else:
        data_start = start_row

    r = data_start
    prev_block: str | None = None

    for key in row_keys:
        if default_row_fill is not None:
            row_fill = default_row_fill
        elif row_fill_fn:
            row_fill = row_fill_fn(key)
        else:
            row_fill = _DATA_ZEBRA if (r - data_start) % 2 else _WHITE
        is_total = row_is_total_fn(key) if row_is_total_fn else False

        # Visual separator between revenue and expense blocks (financial sheets).
        block: str | None = None
        if row_fill_fn:
            if row_fill == _REVENUE_FILL:
                block = "revenue"
            elif row_fill == _EXPENSE_FILL:
                block = "expense"
            if block and prev_block and block != prev_block:
                r += 1  # blank white spacer row
            prev_block = block

        label_cell = ws.cell(row=r, column=1, value=_excel_value(_row_label(label_section, key)))
        _apply_cell_style(
            label_cell,
            kind=CellKind.TEXT,
            row_fill=row_fill,
            bold_label=True,
            is_total_row=is_total,
        )
        for j, col_data in enumerate(columns):
            is_vacant = bool(vacant_columns[j]) if vacant_columns else False
            raw = col_data.get(key)
            if (
                label_section in ("financial_statement", "financial_statement_hotel")
                and key in _FINANCIAL_SUPPRESS_NEGATIVE_OPEX_KEYS
            ):
                nv = _coerce_numeric(raw)
                if nv is not None and nv < 0:
                    raw = None
            display, kind = _prepare_cell(key, raw)
            cell = ws.cell(row=r, column=j + 2, value=_excel_value(display))
            _apply_cell_style(
                cell,
                kind=kind,
                row_fill=row_fill,
                bold_value=is_total,
                is_total_row=is_total,
                vacant=is_vacant,
            )
        r += 1

    return r + 1


def _write_financial_matrix(
    ws: Any,
    start_row: int,
    items: list[Any],
    *,
    fin_key: str,
) -> int:
    return _write_list_matrix(
        ws,
        start_row,
        items,
        label_section=fin_key,
        custom_order=FINANCIAL_ROW_ORDER.get(fin_key),
        row_fill_fn=lambda k: _financial_row_fill(fin_key, k) or _WHITE,
        row_is_total_fn=lambda k: _financial_is_total(fin_key, k),
    )


def _rent_roll_merged_field_order() -> tuple[str, ...]:
    """Single column order: unit / rent / lease blocks concatenated (dedupe)."""
    out: list[str] = []
    seen: set[str] = set()
    for _title, keys in RENT_ROLL_BLOCKS:
        for k in keys:
            if k not in seen:
                seen.add(k)
                out.append(k)
    return tuple(out)


def _rent_roll_table_title(building_name: str | None, building_index: int) -> str:
    if not _is_blank(building_name):
        return str(building_name).strip()
    return f"Building {building_index + 1}"


def _write_rent_roll_transposed_table(
    ws: Any,
    start_row: int,
    units: list[Any],
    *,
    building_name: str | None = None,
    building_index: int = 0,
) -> int:
    """One table per building: title row, header row, then one row per lease (unit fields repeated)."""
    expanded: list[tuple[dict[str, Any], int]] = []
    for unit_idx, raw in enumerate(units):
        if not isinstance(raw, dict):
            continue
        for row in _iter_rent_roll_excel_rows(raw):
            if row and not _dict_all_values_blank(row):
                expanded.append((row, unit_idx))

    flat = [r for r, _ in expanded]
    stripe_by_unit = [s for _, s in expanded]

    base_order = _rent_roll_merged_field_order()
    present_keys = (
        [k for k in base_order if any(not _is_blank(u.get(k)) for u in flat)]
        if flat
        else list(base_order)
    )
    if not present_keys:
        present_keys = ["unit_id", "tenant_name"]

    title = _rent_roll_table_title(building_name, building_index)
    last_col = max(1 + len(present_keys), 2)

    r = start_row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
    title_cell = ws.cell(row=r, column=1, value=_excel_value(title))
    title_cell.font = _SECTION_FONT
    title_cell.fill = _SECTION_FILL
    title_cell.alignment = _LEFT
    title_cell.border = _BORDER
    r += 1

    hdr = r
    _style_table_header(ws, hdr, 1, last_col)
    ws.cell(row=hdr, column=1, value="#")
    for j, key in enumerate(present_keys):
        c = ws.cell(row=hdr, column=j + 2, value=_excel_value(_row_label("rent_roll_report", key)))
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.border = _BORDER
        c.alignment = _CENTER

    r = hdr + 1
    if not flat:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
        empty = ws.cell(row=r, column=1, value="No unit rows extracted for this building.")
        empty.font = _BODY_FONT
        empty.fill = _WHITE
        empty.border = _BORDER
        empty.alignment = _WRAP
        return r + 2

    for i, unit in enumerate(flat):
        vacant = _unit_is_vacant(unit)
        stripe = stripe_by_unit[i]
        row_fill = _DATA_ZEBRA if stripe % 2 else _WHITE
        idx_cell = ws.cell(row=r, column=1, value=i + 1)
        _apply_cell_style(idx_cell, kind=CellKind.INTEGER, row_fill=row_fill, bold_value=False, vacant=vacant)
        for j, key in enumerate(present_keys):
            cell_raw = unit.get(key)
            display, kind = _prepare_cell(key, cell_raw)
            cell = ws.cell(row=r, column=j + 2, value=_excel_value(display))
            _apply_cell_style(cell, kind=kind, row_fill=row_fill, vacant=vacant)
        r += 1

    return r + 1


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------


def _build_summary_sheet(
    ws: Any,
    data: dict[str, Any],
    *,
    source_pdf_path: Path | None = None,
) -> bool:
    _init_sheet_canvas(ws)
    meta = data.get("metadata_from_text") if isinstance(data.get("metadata_from_text"), dict) else {}
    scalars = _metadata_scalar_block(meta)
    offers = _metadata_offer_rows(meta)
    cap_d, cap_k, noi_d, noi_k, year_lbl = _first_financial_year_kpis(data)

    has_profile = _has_meaningful_value(scalars) or _has_meaningful_value(offers)
    fin_key = _pick_financial_key(data)
    fin_section = data.get(fin_key) if fin_key else None
    cycles = fin_section.get("financial_cycles") if isinstance(fin_section, dict) else None
    has_first_year = isinstance(cycles, list) and bool(cycles)

    if not has_profile and not has_first_year:
        return False

    offer_name = str(meta.get("offer_name") or "").strip()
    loc_bits = [str(meta.get(k) or "").strip() for k in ("city", "state") if not _is_blank(meta.get(k))]
    subtitle = " — ".join(p for p in (offer_name, ", ".join(loc_bits)) if p) or None

    row = 1
    row = _write_hero_header(ws, row, "Overall offer description", subtitle, 6)
    if has_first_year:
        row = _write_financial_kpi_strip(
            ws,
            row,
            year_label=year_lbl,
            cap_val=cap_d,
            cap_kind=cap_k,
            noi_val=noi_d,
            noi_kind=noi_k,
        )
    row = _write_page_block(
        ws,
        row,
        _pages_for_sheet(data, "summary", meta),
        confidence_section=meta,
    )
    if scalars:
        row = _write_section_banner(ws, row, "Deal profile", 2)
        row = _write_single_column_table(ws, row, scalars, label_section="metadata_from_text")
    if offers:
        row = _write_offer_lines_matrix(ws, row, offers, data=data)
    _autosize_columns(ws, max(ws.max_column, 6))
    _embed_excel_sheet_key_pages_below(ws, row, data, "summary", source_pdf_path)
    return True


def _build_financial_sheet(
    ws: Any,
    data: dict[str, Any],
    *,
    source_pdf_path: Path | None = None,
) -> bool:
    _init_sheet_canvas(ws)
    fin_key = _pick_financial_key(data)
    if not fin_key:
        return False
    section = data.get(fin_key)
    if not isinstance(section, dict):
        return False
    years = section.get("financial_cycles")
    if not isinstance(years, list) or not _has_meaningful_value(years):
        return False

    ncol = 1 + len(_filter_list_columns(years))
    row = 1
    row = _write_hero_header(ws, row, "Financial statement", "Operating statement by period", max(ncol, 4))
    row = _write_page_block(
        ws,
        row,
        _pages_for_sheet(data, "financial_statement", section),
        confidence_section=section,
    )
    row = _write_section_banner(ws, row, "Line items", max(ncol, 2))
    row = _write_financial_matrix(ws, row, years, fin_key=fin_key)
    _autosize_columns(ws, max(ws.max_column, 2))
    _embed_excel_sheet_key_pages_below(ws, row, data, "financial_statement", source_pdf_path)
    return True


def _build_rent_roll_sheet(
    ws: Any,
    data: dict[str, Any],
    *,
    source_pdf_path: Path | None = None,
) -> bool:
    _init_sheet_canvas(ws)
    rr = data.get("rent_roll_report")
    if not isinstance(rr, dict):
        return False
    sections = _rent_roll_building_sections(rr)
    if not sections:
        return False

    nkeys = len(_rent_roll_merged_field_order())
    ncol = max(6, 1 + nkeys)

    row = 1
    row = _write_hero_header(ws, row, "Rent roll", "One table per building; one row per lease term when listed", max(ncol, 4))
    row = _write_page_block(
        ws,
        row,
        _pages_for_sheet(data, "rent_roll", rr),
        confidence_section=rr,
    )
    for bidx, (bname, units) in enumerate(sections):
        if bidx:
            row += 2
        row = _write_rent_roll_transposed_table(
            ws,
            row,
            units,
            building_name=bname,
            building_index=bidx,
        )
    _autosize_columns(ws, max(ws.max_column, 2))
    _embed_excel_sheet_key_pages_below(ws, row, data, "rent_roll", source_pdf_path)
    return True


def _build_building_sheet(
    ws: Any,
    data: dict[str, Any],
    *,
    source_pdf_path: Path | None = None,
) -> bool:
    _init_sheet_canvas(ws)
    building_raw = data.get("building_report") if isinstance(data.get("building_report"), dict) else {}
    raw_assets = building_raw.get("assets")
    assets = [x for x in raw_assets if isinstance(x, dict)] if isinstance(raw_assets, list) else []
    if not _has_meaningful_value(assets):
        return False

    ncol = 1 + max(len(_filter_list_columns(assets)), 1)
    row = 1
    row = _write_hero_header(ws, row, "Building report", "Physical asset and technical detail", max(ncol, 4))
    row = _write_page_block(
        ws,
        row,
        _pages_for_sheet(data, "building_report", building_raw),
        confidence_section=building_raw,
    )
    row = _write_section_banner(ws, row, "Buildings", max(ncol, 2))
    row = _write_list_matrix(ws, row, assets, label_section="building_report")
    _autosize_columns(ws, max(ws.max_column, 2))
    _embed_excel_sheet_key_pages_below(ws, row, data, "building_report", source_pdf_path)
    return True


def _build_demographics_sheet(
    ws: Any,
    data: dict[str, Any],
    *,
    source_pdf_path: Path | None = None,
) -> bool:
    _init_sheet_canvas(ws)
    demo = data.get("demographics_report") if isinstance(data.get("demographics_report"), dict) else {}
    cols = demo.get("area_statistics")
    if not isinstance(cols, list) or not cols:
        legacy = demo.get("catchment_areas")
        cols = legacy if isinstance(legacy, list) else []
    if not isinstance(cols, list) or not _has_meaningful_value(cols):
        return False

    ncol = 1 + len(_filter_list_columns([x for x in cols if isinstance(x, dict)]))
    row = 1
    row = _write_hero_header(ws, row, "Demographics report", "Market depth, income, and composition", max(ncol, 4))
    row = _write_page_block(
        ws,
        row,
        _pages_for_sheet(data, "demographics_report", demo),
        confidence_section=demo,
    )
    row = _write_section_banner(ws, row, "Catchment statistics", max(ncol, 2))
    row = _write_list_matrix(ws, row, cols, label_section="demographics_report")
    _autosize_columns(ws, max(ws.max_column, 2))
    _embed_excel_sheet_key_pages_below(ws, row, data, "demographics_report", source_pdf_path)
    return True


def _build_offer_pictures_sheet(
    ws: Any,
    data: dict[str, Any],
    *,
    source_pdf_path: Path | None = None,
) -> bool:
    """Embed PDF page screenshots for ``property_pictures_page`` (PageOfInterest)."""
    pages = _page_list_for_poi_fields(data, "property_pictures_page")
    if not pages:
        return False
    if source_pdf_path is None or not Path(source_pdf_path).is_file():
        logger.warning("offer_pictures sheet skipped: PDF path missing or not found.")
        return False

    pdf_path = Path(source_pdf_path).resolve()
    _init_sheet_canvas(ws)
    ws.column_dimensions["A"].width = max(100.0, min(130.0, 24.0 + float(_EMBED_MAX_WIDTH_PX) / 85.0))

    row = 1
    row = _write_hero_header(ws, row, "Offer pictures", "Property photos from the offering memorandum", 4)
    page_ex = data.get("page_extraction") if isinstance(data.get("page_extraction"), dict) else None
    row = _write_page_block(
        ws,
        row,
        _format_page_list(pages),
        confidence_section=page_ex,
    )

    embedded = 0
    z = float(_EMBED_RENDER_ZOOM)
    cap = int(_EMBED_MAX_WIDTH_PX)
    for page_num in pages:
        png = _render_pdf_page_png(pdf_path, page_num, zoom=z)
        if not png:
            continue
        xl_img = _scaled_xl_image(png, max_width_px=cap)
        if xl_img is None:
            continue

        ws.cell(row=row, column=1, value=f"Page {page_num}").font = _SECTION_FONT
        ws.cell(row=row, column=1).fill = _SECTION_FILL
        ws.cell(row=row, column=1).border = _BORDER
        row += 1

        anchor = f"A{row}"
        ws.add_image(xl_img, anchor)
        row += _anchor_image_rows(int(xl_img.height))
        embedded += 1
        row += 1

    if embedded == 0:
        ws.cell(row=row, column=1, value="Could not render property picture pages from the PDF.").font = _BODY_FONT
        return False

    return True


_SHEET_BUILDERS = {
    "summary": _build_summary_sheet,
    "offer_pictures": _build_offer_pictures_sheet,
    "rent_roll": _build_rent_roll_sheet,
    "financial_statement": _build_financial_sheet,
    "building_report": _build_building_sheet,
    "demographics_report": _build_demographics_sheet,
}


def build_text_llm_workbook(
    text_llm_by_schema: Optional[dict[str, Any]],
    *,
    source_pdf_path: Path | None = None,
) -> Any:
    if Workbook is None:
        raise ImportError("openpyxl is required — pip install openpyxl")

    data = text_llm_by_schema or {}
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name in EXCEL_WORKBOOK_SHEETS:
        builder = _SHEET_BUILDERS.get(sheet_name)
        if builder is None:
            continue
        ws = wb.create_sheet(title=sheet_name[:31])
        ok = builder(ws, data, source_pdf_path=source_pdf_path)
        if not ok:
            wb.remove(ws)

    if not wb.sheetnames:
        ws = wb.create_sheet(title="summary")
        _init_sheet_canvas(ws)
        ws.cell(row=1, column=1, value="No extraction data available.")

    for sheet in wb.worksheets:
        sheet.freeze_panes = None

    return wb


def write_text_llm_excel(
    text_llm_by_schema: Optional[dict[str, Any]],
    output_path: Path,
    *,
    source_pdf_path: Path | None = None,
) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = build_text_llm_workbook(text_llm_by_schema, source_pdf_path=source_pdf_path)
    wb.save(str(path))
    return path.resolve()


def save_parse_excel_export(
    *,
    text_llm_by_schema: Optional[dict[str, Any]],
    source_pdf_path: Path,
) -> Optional[Path]:
    try:
        out = source_pdf_path.resolve().parent / f"{source_pdf_path.stem}_extraction.xlsx"
        return write_text_llm_excel(text_llm_by_schema, out, source_pdf_path=source_pdf_path)
    except ImportError:
        logger.warning("openpyxl not installed — skipping Excel export.")
        return None
    except Exception as exc:
        logger.exception("Excel export failed: %s", exc)
        return None
